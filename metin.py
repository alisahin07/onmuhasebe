from openpyxl import Workbook, load_workbook
def gonder():

    wb = load_workbook (filename='demoexcel.xlsx')
    sh = wb['Sheet']
    row_ct = sh.max_row
    col_ct = sh.max_column
    print(row_ct)
    print(col_ct)
    dizi=[]
    for i in range(1, row_ct + 1):
        for j in range(1, col_ct ):
            if sh.cell(row=i, column=j+1).value!=None:
             restoran_name=sh.cell(row=i, column=j).value
             telefon_number=sh.cell(row=i, column=j+1).value
             print(restoran_name+'--'+telefon_number)
             dizi.append([restoran_name,telefon_number])
    return dizi


